from decimal import Decimal
import requests
import os
import stripe
from django.db.models import Count
import json
from django.conf import settings
from django.http import FileResponse, Http404
from rest_framework import viewsets, permissions, serializers
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from django.utils import timezone
from datetime import timedelta
from .permissions import IsAdminOrReadOnly
from .permissions import IsAdmin
from .services.supabase_storage import SupabaseStorageService
import logging
from rest_framework.parsers import MultiPartParser, FormParser
import traceback
from datetime import date
from django.db import models
from django.db import IntegrityError, transaction
from rest_framework import status
from rest_framework.response import Response
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from datetime import datetime, timedelta
from django.db.models import Q, Sum
from .models import Bitacora, Usuario
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

AI_WORKER_URL = os.getenv("AI_WORKER_URL")
from .models import (
    Rol, Usuario, Propiedad, Multa, Pagos, Notificaciones, AreasComunes, Tareas,
    Vehiculo, Pertenece, ListaVisitantes, DetalleMulta, Factura, Finanzas,
    Comunicados, Horarios, Reserva, Asignacion, Envio, Registro, Bitacora,
    PerfilFacial, ReconocimientoFacial, DeteccionPlaca, ReporteSeguridad, SolicitudMantenimiento,
    MantenimientoPreventivo
)
from .serializers import (
    RolSerializer, UsuarioSerializer, PropiedadSerializer, MultaSerializer,
    PagoSerializer, NotificacionesSerializer, AreasComunesSerializer, TareasSerializer,
    VehiculoSerializer, PerteneceSerializer, ListaVisitantesSerializer, DetalleMultaSerializer,
    FacturaSerializer, FinanzasSerializer, ComunicadosSerializer, HorariosSerializer,
    ReservaSerializer, AsignacionSerializer, EnvioSerializer, RegistroSerializer,
    BitacoraSerializer, ReconocimientoFacialSerializer, PerfilFacialSerializer, DeteccionPlacaSerializer,
    ReporteSeguridadSerializer, EstadoCuentaSerializer, PagoRealizadoSerializer, SolicitudMantenimientoSerializer,
    MantenimientoPreventivoSerializer
)


# ---------------------------------------------------------------------
# Base genérica: agrega filtros, búsqueda y ordenamiento a todos
# ---------------------------------------------------------------------
class BaseModelViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    # Cada viewset define: queryset, serializer_class, filterset_fields, search_fields, ordering_fields


# ---------------------------------------------------------------------
# Catálogos / tablas simples
# ---------------------------------------------------------------------
class RolViewSet(BaseModelViewSet):
    queryset = Rol.objects.all().order_by('id')
    serializer_class = RolSerializer
    filterset_fields = ['tipo', 'estado']
    search_fields = ['descripcion', 'tipo', 'estado']
    ordering_fields = ['id', 'tipo', 'estado']


class PropiedadViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]
    queryset = Propiedad.objects.all().order_by('nro_casa', 'piso')
    serializer_class = PropiedadSerializer
    filterset_fields = ['nro_casa', 'piso', 'descripcion']
    search_fields = ['descripcion', 'nro_casa']
    ordering_fields = ['codigo', 'nro_casa', 'piso', 'tamano_m2']

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.query_params.get('include_residents') == 'true':
            queryset = queryset.prefetch_related('pertenentes__codigo_usuario')
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)

        propiedades_data = self._serialize_propiedades_with_residents(page if page is not None else queryset)

        if page is not None:
            return self.get_paginated_response(propiedades_data)
        return Response(propiedades_data)

    def create(self, request, *args, **kwargs):
        nro_casa = request.data.get('nro_casa')
        piso = request.data.get('piso', 0)
        if nro_casa is not None and piso is not None:
            if Propiedad.objects.filter(nro_casa=nro_casa, piso=piso).exists():
                return Response(
                    {'detail': f'Ya existe una unidad con número {nro_casa} en piso {piso}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        try:
            usuario = Usuario.objects.get(correo=request.user.email)
            Bitacora.objects.create(
                codigo_usuario=usuario,
                accion=f"Creación de unidad habitacional {nro_casa}",
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                ip=self._get_client_ip(request)
            )
        except Usuario.DoesNotExist:
            pass
        try:
            with transaction.atomic():
                return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {"detail": "La unidad ya existe (NroCasa + Piso)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        nro_casa = request.data.get("nro_casa", instance.nro_casa)
        piso = request.data.get("piso", instance.piso if instance.piso is not None else 0)
        if (
                nro_casa is not None
                and piso is not None
                and Propiedad.objects.filter(nro_casa=nro_casa, piso=piso)
                .exclude(pk=instance.pk)
                .exists()
        ):
            return Response(
                {"detail": f"Ya existe una unidad con número {nro_casa} en piso {piso}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            usuario = Usuario.objects.get(correo=request.user.email)
            Bitacora.objects.create(
                codigo_usuario=usuario,
                accion=f"Edición de unidad habitacional {instance.nro_casa}",
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                ip=self._get_client_ip(request)
            )
        except Usuario.DoesNotExist:
            pass
        try:
            with transaction.atomic():
                return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {"detail": "La unidad ya existe (NroCasa + Piso)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # --- INICIO DE LA CORRECCIÓN DE INDENTACIÓN ---
    # Estas funciones ahora están al nivel correcto de la clase.
    def _serialize_propiedades_with_residents(self, propiedades):
        """Serializa propiedades incluyendo una lista de residentes actuales y sus roles."""
        result = []
        # --- ESTE ES EL CAMBIO CLAVE ---
        # Usamos timezone.now().date() en lugar de date.today()
        # para asegurarnos de usar la zona horaria de settings.py
        today = timezone.now().date()

        for propiedad in propiedades:
            prop_data = PropiedadSerializer(propiedad).data

            # El resto de la lógica no necesita cambios
            vinculaciones_activas = [
                p for p in propiedad.pertenentes.all()
                if p.fecha_ini <= today and (p.fecha_fin is None or p.fecha_fin >= today)
            ]

            residentes_actuales = []
            for v in vinculaciones_activas:
                if v.codigo_usuario:
                    residentes_actuales.append({
                        'pertenece_id': v.id,
                        'codigo_usuario': v.codigo_usuario.codigo,
                        'nombre_completo': f"{v.codigo_usuario.nombre} {v.codigo_usuario.apellido}",
                        'correo': v.codigo_usuario.correo,
                        'rol_en_propiedad': v.rol_en_propiedad,
                        'fecha_ini': v.fecha_ini,
                        'fecha_fin': v.fecha_fin,
                    })

            prop_data['residentes_actuales'] = residentes_actuales
            result.append(prop_data)

        return result

    def _get_client_ip(self, request):
        """Obtiene la IP del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    # --- FIN DE LA CORRECCIÓN DE INDENTACIÓN ---


class BitacoraMixin:
    def _bitacora(self, request, accion: str):
        try:
            usuario = Usuario.objects.get(correo=getattr(request.user, "email", None))
            Bitacora.objects.create(
                codigo_usuario=usuario,
                accion=accion,
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                ip=self._get_client_ip(request),
            )
        except Usuario.DoesNotExist:
            pass

    def _get_client_ip(self, request):
        xf = request.META.get("HTTP_X_FORWARDED_FOR")
        return xf.split(",")[0] if xf else request.META.get("REMOTE_ADDR")


class MultaViewSet(BitacoraMixin, viewsets.ModelViewSet):
    permission_classes = [IsAdmin]
    queryset = Multa.objects.all().order_by("descripcion")
    serializer_class = MultaSerializer
    search_fields = ["descripcion"]
    filterset_fields = (["estado"] if hasattr(Multa, "estado") else [])
    ordering_fields = ["id", "descripcion", "monto"]

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                resp = super().create(request, *args, **kwargs)
                self._bitacora(request, f"Alta de multa: {resp.data.get('descripcion')}")
                return resp
        except IntegrityError:
            # Por si luego agregas UNIQUE en descripcion
            return Response({"detail": "Ya existe una multa con esa descripción."}, status=400)

    def update(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                resp = super().update(request, *args, **kwargs)
                self._bitacora(request, f"Edición de multa: {resp.data.get('descripcion')}")
                return resp
        except IntegrityError:
            return Response({"detail": "Conflicto de BD al actualizar la multa."}, status=400)


class PagoViewSet(BitacoraMixin, viewsets.ModelViewSet):
    permission_classes = [IsAdminOrReadOnly]
    queryset = Pagos.objects.all().order_by("tipo", "descripcion")
    serializer_class = PagoSerializer
    search_fields = ["tipo", "descripcion"]
    filterset_fields = ["tipo"] + (["estado"] if hasattr(Pagos, "estado") else [])
    ordering_fields = ["id", "tipo", "descripcion", "monto"]

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                resp = super().create(request, *args, **kwargs)
                self._bitacora(request, f"Alta de pago: {resp.data.get('descripcion')}")
                return resp
        except IntegrityError:
            return Response({"detail": "No se pudo crear el pago (conflicto en BD)."}, status=400)

    def update(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                resp = super().update(request, *args, **kwargs)
                self._bitacora(request, f"Edición de pago: {resp.data.get('descripcion')}")
                return resp
        except IntegrityError:
            return Response({"detail": "No se pudo actualizar el pago (conflicto en BD)."}, status=400)


class NotificacionesViewSet(BaseModelViewSet):
    queryset = Notificaciones.objects.all().order_by('id')
    serializer_class = NotificacionesSerializer
    filterset_fields = ['tipo']
    search_fields = ['tipo', 'descripcion']
    ordering_fields = ['id']


class AreasComunesViewSet(BaseModelViewSet):
    permission_classes = [IsAdminOrReadOnly]
    queryset = AreasComunes.objects.all().order_by('id')
    serializer_class = AreasComunesSerializer
    filterset_fields = ['estado', 'capacidad_max', 'costo']  # <-- Corregido
    search_fields = ['descripcion', 'estado']
    ordering_fields = ['id', 'capacidad_max', 'costo']  # <-- Corregido

    @action(detail=True, methods=['get'])
    def disponibilidad(self, request, pk=None):
        """
        Endpoint para obtener los horarios disponibles de un área común en una fecha específica.
        Uso: /api/areas-comunes/1/disponibilidad/?fecha=2025-12-31
        """
        try:
            fecha_str = request.query_params.get('fecha')
            if not fecha_str:
                return Response({'error': 'Debe proporcionar una fecha en formato YYYY-MM-DD.'}, status=400)

            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            area = self.get_object()

            # --- LÓGICA DE DISPONIBILIDAD MEJORADA ---
            # Un área solo está ocupada si hay una reserva en estado 'Aprobada'.
            reservas_aprobadas = Reserva.objects.filter(idareac=area, fecha=fecha_obj, estado='Aprobada')

            disponible = not reservas_aprobadas.exists()

            return Response({
                'area_id': area.id,
                'area_descripcion': area.descripcion,
                'fecha_consultada': fecha_obj.isoformat(),
                'disponible': disponible,
            })

        except ValueError:
            return Response({'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)
        except AreasComunes.DoesNotExist:
            return Response({'error': 'Área común no encontrada.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TareasViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]
    queryset = Tareas.objects.all().order_by('id')
    serializer_class = TareasSerializer
    filterset_fields = ['tipo', 'vigencia', 'costos']
    search_fields = ['tipo', 'descripcion']
    ordering_fields = ['id', 'vigencia', 'costos']


class VehiculoViewSet(BaseModelViewSet):
    serializer_class = VehiculoSerializer
    filterset_fields = ['estado', 'nro_placa', 'codigo_usuario']  # <-- CAMBIADO
    search_fields = ['nro_placa', 'descripcion', 'estado']
    ordering_fields = ['id']

    def get_queryset(self):
        try:
            usuario = Usuario.objects.get(correo=self.request.user.email)
            if usuario.idrol and usuario.idrol.tipo == 'admin':
                return Vehiculo.objects.all().order_by('id')
            return Vehiculo.objects.filter(codigo_usuario=usuario).order_by('id')
        except Usuario.DoesNotExist:
            return Vehiculo.objects.none()

    def perform_create(self, serializer):
        try:
            usuario = Usuario.objects.get(correo=self.request.user.email)
            serializer.save(codigo_usuario=usuario, estado='activo')
        except Usuario.DoesNotExist:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Tu usuario no está registrado en el catálogo para realizar esta acción.")


# ---------------------------------------------------------------------
# Entidades con FK
# ---------------------------------------------------------------------
class UsuarioViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]  # <-- 1. SOLO ADMINS PUEDEN GESTIONAR USUARIOS
    queryset = Usuario.objects.all().order_by('codigo')
    serializer_class = UsuarioSerializer
    filterset_fields = ['idrol', 'sexo', 'estado', 'correo', 'telefono']
    search_fields = ['nombre', 'apellido', 'correo', 'estado']
    ordering_fields = ['codigo', 'telefono']

    def destroy(self, request, *args, **kwargs):
        """
        Sobrescribe el método de borrado para asegurar una eliminación completa.
        """
        instance = self.get_object()

        # 2. Busca y elimina el usuario de autenticación de Django
        # para prevenir cuentas huérfanas.
        dj_user = User.objects.filter(username=instance.correo).first()
        if dj_user:
            dj_user.delete()

        # 3. Llama al método de borrado original para eliminar el registro de la tabla Usuario.
        self.perform_destroy(instance)

        # Devuelve una respuesta de éxito sin contenido.
        return Response(status=status.HTTP_204_NO_CONTENT)


# smartcondominium-back/api/views.py

# ... (otros imports)

# REEMPLAZA TU CLASE PerteneceViewSet ACTUAL CON ESTA VERSIÓN MEJORADA
class PerteneceViewSet(BaseModelViewSet):
    # Ya no usamos permission_classes aquí para poder definir permisos por acción
    serializer_class = PerteneceSerializer
    queryset = Pertenece.objects.all().order_by('-fecha_ini')
    filterset_fields = ['codigo_usuario', 'codigo_propiedad', 'fecha_ini', 'fecha_fin', 'rol_en_propiedad']

    def get_permissions(self):
        """
        Instancia y retorna la lista de permisos que esta vista requiere.
        """
        if self.action in ['list', 'retrieve']:
            # Para 'ver' la lista o un detalle, solo se necesita estar autenticado.
            permission_classes = [IsAuthenticated]
        else:
            # Para cualquier otra acción (crear, editar, borrar), se necesita ser admin.
            permission_classes = [IsAdmin]
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        """
        Filtra el queryset para que los residentes solo vean sus propias asignaciones.
        Los administradores ven todo.
        """
        user = self.request.user
        queryset = super().get_queryset()

        try:
            usuario_actual = Usuario.objects.get(correo=user.email)
            if not (usuario_actual.idrol and usuario_actual.idrol.tipo == 'admin'):
                # Si no es admin, filtramos por su propio código de usuario.
                queryset = queryset.filter(codigo_usuario=usuario_actual)

            # Implementamos el filtro personalizado 'activas' que pide el frontend
            if self.request.query_params.get('activas') == 'true':
                from django.utils import timezone
                today = timezone.now().date()
                queryset = queryset.filter(
                    fecha_ini__lte=today,
                ).filter(
                    models.Q(fecha_fin__isnull=True) | models.Q(fecha_fin__gte=today)
                )

        except Usuario.DoesNotExist:
            return queryset.none()

        return queryset

    # El resto de los métodos de PerteneceViewSet (create, destroy, etc.) no necesitan cambios.
    # ... (pega aquí tus métodos create, destroy y _get_client_ip que ya tenías)
    def create(self, request, *args, **kwargs):
        codigo_usuario = request.data.get('codigo_usuario')
        codigo_propiedad = request.data.get('codigo_propiedad')
        fecha_ini = request.data.get('fecha_ini')
        fecha_fin = request.data.get('fecha_fin')
        rol_asignado = request.data.get('rol_en_propiedad')

        if not all([codigo_usuario, codigo_propiedad, fecha_ini, rol_asignado]):
            return Response(
                {'detail': 'Usuario, propiedad, fecha de inicio y rol son obligatorios.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            usuario = Usuario.objects.get(codigo=codigo_usuario, estado='activo')
            propiedad = Propiedad.objects.get(codigo=codigo_propiedad)
        except Usuario.DoesNotExist:
            return Response({'detail': 'Usuario no encontrado o inactivo.'}, status=status.HTTP_404_NOT_FOUND)
        except Propiedad.DoesNotExist:
            return Response({'detail': 'Propiedad no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        from datetime import datetime
        fecha_ini_date = datetime.strptime(fecha_ini, '%Y-%m-%d').date()

        # Regla estricta: solo permite UNA propiedad activa por usuario
        conflicto_usuario = Pertenece.objects.filter(
            codigo_usuario=usuario,
            fecha_ini__lte=fecha_ini_date,
        ).filter(
            models.Q(fecha_fin__isnull=True) | models.Q(fecha_fin__gte=fecha_ini_date)
        ).exists()

        if conflicto_usuario:
            return Response(
                {'detail': 'Este usuario ya está vinculado a OTRA propiedad en el período especificado.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if rol_asignado == 'Propietario':
            conflicto_propietario = Pertenece.objects.filter(
                codigo_propiedad=propiedad,
                rol_en_propiedad='Propietario',
                fecha_ini__lte=fecha_ini_date,
            ).filter(
                models.Q(fecha_fin__isnull=True) | models.Q(fecha_fin__gte=fecha_ini_date)
            ).exists()

            if conflicto_propietario:
                return Response(
                    {'detail': 'Esta propiedad ya tiene un Propietario asignado en el período especificado.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            admin_usuario = Usuario.objects.get(correo=request.user.email)
            Bitacora.objects.create(
                codigo_usuario=admin_usuario,
                accion=f"Vinculación de {usuario.nombre} a unidad {propiedad.nro_casa} como {rol_asignado}",
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                ip=self._get_client_ip(request)
            )
        except Exception:
            pass

        return super().create(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        usuario_asignado = instance.codigo_usuario
        propiedad_asignada = instance.codigo_propiedad
        try:
            admin_usuario = Usuario.objects.get(correo=request.user.email)
            Bitacora.objects.create(
                codigo_usuario=admin_usuario,
                accion=f"Desvinculación de {usuario_asignado.nombre} de la unidad {propiedad_asignada.nro_casa}",
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                ip=self._get_client_ip(request)
            )
        except Exception:
            pass
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def _get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class ListaVisitantesViewSet(BaseModelViewSet):
    queryset = ListaVisitantes.objects.all().order_by('id')
    serializer_class = ListaVisitantesSerializer
    filterset_fields = ['codigo_propiedad', 'fecha_ini', 'fecha_fin', 'carnet']
    search_fields = ['nombre', 'apellido', 'carnet', 'motivovisita']
    ordering_fields = ['id', 'fechaini', 'fechafin']

    @action(detail=True, methods=['get'])
    def generar_pase(self, request, pk=None):
        """
        Genera una URL para el pase de acceso QR que lleva a una página pública.
        """
        try:
            visitante = self.get_object()
            propiedad = visitante.codigo_propiedad

            pass_data = {
                "type": "ACCESS_PASS",
                "visitante_id": visitante.id,
                "nombre": f"{visitante.nombre} {visitante.apellido}",
                "carnet": visitante.carnet,
                "propiedad_destino": f"Casa {propiedad.nro_casa}, Piso {propiedad.piso}",
                "valido_desde": visitante.fecha_ini.isoformat(),
                "valido_hasta": visitante.fecha_fin.isoformat(),
            }

            # 1. Convertimos el diccionario a un string JSON
            json_string = json.dumps(pass_data, ensure_ascii=False)

            # 2. Codificamos el string JSON en Base64 para que sea seguro en una URL
            import base64
            base64_string = base64.b64encode(json_string.encode('utf-8')).decode('utf-8')

            # 3. Obtenemos la URL del frontend desde la configuración
            from django.conf import settings
            frontend_url = settings.FRONTEND_URL

            # 4. Creamos la URL completa que irá en el QR
            pase_url = f"{frontend_url}/pase-visitante?data={base64_string}"

            # 5. Enviamos la URL al frontend
            return Response({'qr_data': pase_url}, status=status.HTTP_200_OK)

        except ListaVisitantes.DoesNotExist:
            return Response({'error': 'Visitante no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class DetalleMultaViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]
    queryset = DetalleMulta.objects.all().order_by('id')
    serializer_class = DetalleMultaSerializer
    filterset_fields = ['codigo_propiedad', 'id_multa', 'fecha_emi', 'fecha_lim']
    search_fields = []
    ordering_fields = ['id', 'fecha_emi', 'fecha_lim']


class FacturaViewSet(BaseModelViewSet):
    queryset = Factura.objects.all().order_by('id')
    serializer_class = FacturaSerializer
    filterset_fields = ['codigousuario', 'idpago', 'fecha', 'estado', 'tipopago']
    search_fields = ['estado', 'tipopago']
    ordering_fields = ['id', 'fecha']


class FinanzasViewSet(BaseModelViewSet):
    queryset = Finanzas.objects.all().order_by('id')
    serializer_class = FinanzasSerializer
    filterset_fields = ['tipo', 'fecha', 'origen', 'idfactura']
    search_fields = ['tipo', 'descripcion', 'origen']
    ordering_fields = ['id', 'fecha', 'monto']


# En api/views.py
class ComunicadosViewSet(BaseModelViewSet):
    queryset = Comunicados.objects.all().order_by('id')
    serializer_class = ComunicadosSerializer
    permission_classes = [IsAdminOrReadOnly]  # <-- 2. REEMPLAZA EL PERMISO AQUÍ
    filterset_fields = ['tipo', 'fecha', 'estado']
    search_fields = ['titulo', 'contenido', 'url', 'tipo', 'estado']
    ordering_fields = ['id', 'fecha']


class HorariosViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]
    queryset = Horarios.objects.all().order_by('id')
    serializer_class = HorariosSerializer
    filterset_fields = ['id_area_c', 'hora_ini', 'hora_fin']
    search_fields = []
    ordering_fields = ['id', 'hora_ini', 'hora_fin']


class ReservaViewSet(BaseModelViewSet):
    queryset = Reserva.objects.all().order_by('-fecha')
    serializer_class = ReservaSerializer
    # CORREGIDO: Nombres de campos con guion bajo
    filterset_fields = ['codigo_usuario', 'id_area_c', 'fecha', 'estado']
    search_fields = ['estado']
    ordering_fields = ['id', 'fecha']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        try:
            usuario_actual = Usuario.objects.get(correo=user.email)
            if usuario_actual.idrol and usuario_actual.idrol.tipo != 'admin':
                # CORREGIDO: 'codigo_usuario' con guion bajo
                queryset = queryset.filter(codigo_usuario=usuario_actual)
        except Usuario.DoesNotExist:
            return queryset.none()
        return queryset

    def perform_create(self, serializer):
        """
        Asigna automáticamente el usuario actual a la reserva al momento de crearla.
        """
        try:
            # Busca nuestro modelo 'Usuario' a partir del usuario autenticado
            usuario_actual = Usuario.objects.get(correo=self.request.user.email)
            # Guarda la reserva asociándola con este usuario
            serializer.save(codigo_usuario=usuario_actual)
        except Usuario.DoesNotExist:
            # Esto previene errores si el usuario no está en nuestro catálogo
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("El usuario no existe en el sistema de condominio.")

    def create(self, request, *args, **kwargs):
        id_area_c_val = request.data.get('id_area_c')
        fecha_str = request.data.get('fecha')

        if not id_area_c_val or not fecha_str:
            return Response({'detail': 'Se requiere el área común y la fecha.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'El formato de fecha debe ser YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        conflicto = Reserva.objects.filter(id_area_c_id=id_area_c_val, fecha=fecha_obj, estado='Aprobada').exists()
        if conflicto:
            return Response({'detail': 'Esta área ya tiene una reserva aprobada para la fecha seleccionada.'},
                            status=status.HTTP_409_CONFLICT)

        # Hacemos una copia mutable para poder modificarla
        mutable_data = request.data.copy()
        mutable_data['estado'] = 'Pendiente'

        # Pasamos los datos modificados para la validación y creación
        serializer = self.get_serializer(data=mutable_data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin])
    def update_status(self, request, pk=None):
        reserva = self.get_object()
        nuevo_estado = request.data.get('estado')

        if nuevo_estado not in ['Aprobada', 'Rechazada']:
            return Response({'error': 'El estado solo puede ser "Aprobada" o "Rechazada".'},
                            status=status.HTTP_400_BAD_REQUEST)

        if nuevo_estado == 'Aprobada':
            conflicto = Reserva.objects.filter(
                id_area_c=reserva.id_area_c,
                fecha=reserva.fecha,
                estado='Aprobada'
            ).exclude(pk=pk).exists()

            if conflicto:
                return Response(
                    {'error': 'No se puede aprobar. Ya existe otra reserva aprobada para esta fecha y área.'},
                    status=status.HTTP_409_CONFLICT)

        reserva.estado = nuevo_estado
        reserva.save()
        serializer = self.get_serializer(reserva)
        return Response(serializer.data)


class AsignacionViewSet(BaseModelViewSet):
    permission_classes = [IsAdmin]
    queryset = Asignacion.objects.all().order_by('id')
    serializer_class = AsignacionSerializer
    filterset_fields = ['codigo_usuario', 'id_tarea', 'fecha_ini', 'fecha_fin', 'estado']
    search_fields = ['descripcion', 'dificultades', 'estado']
    ordering_fields = ['id', 'fechaini', 'fechafin', 'costo']


class EnvioViewSet(BaseModelViewSet):
    queryset = Envio.objects.all().order_by('id')
    serializer_class = EnvioSerializer
    filterset_fields = ['codigousuario', 'idnotific', 'fecha', 'estado']
    search_fields = ['estado']
    ordering_fields = ['id', 'fecha']


class RegistroViewSet(BaseModelViewSet):
    queryset = Registro.objects.all().order_by('id')
    serializer_class = RegistroSerializer
    filterset_fields = ['codigousuario', 'idvehic', 'fecha']
    search_fields = []
    ordering_fields = ['id', 'fecha', 'hora']


class BitacoraViewSet(BaseModelViewSet):
    queryset = Bitacora.objects.all().order_by('-fecha', '-hora')
    serializer_class = BitacoraSerializer
    filterset_fields = ['codigousuario', 'fecha', 'accion', 'ip']
    search_fields = ['accion', 'ip']
    ordering_fields = ['id', 'fecha', 'hora']


# CU01. Iniciar sesion
class LoginView(APIView):
    permission_classes = [permissions.AllowAny]  # sin auth para poder loguear

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response({"detail": "email y password son requeridos."},
                            status=status.HTTP_400_BAD_REQUEST)

        # 1) Buscar en TU tabla Usuario
        try:
            u = Usuario.objects.get(correo=email)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no existe."}, status=status.HTTP_404_NOT_FOUND)

            # 2) Comparación simple (demo). En prod: NO uses texto plano.
        if u.contrasena != password:
            return Response({"detail": "Credenciales inválidas."}, status=status.HTTP_401_UNAUTHORIZED)

        # 2) Comparación simple (demo). En prod: NO uses texto plano.
        if u.contrasena != password:
            return Response({"detail": "Credenciales inválidas."}, status=status.HTTP_401_UNAUTHORIZED)

        # 3) Sincronizar/crear auth.User para usar TokenAuth
        dj_user, _ = User.objects.get_or_create(username=email, defaults={"email": email})
        if not dj_user.has_usable_password() or not dj_user.check_password(password):
            dj_user.set_password(password)
            dj_user.save()

        # 4) Crear/obtener token
        token, _ = Token.objects.get_or_create(user=dj_user)

        # 5) Armar payload con datos del usuario (y rol)
        rol_obj = None
        if u.idrol_id:
            try:
                r = Rol.objects.get(pk=u.idrol_id)
                rol_obj = {
                    "id": r.id,
                    "descripcion": r.descripcion,
                    "tipo": r.tipo,
                    "estado": r.estado,
                }
            except Rol.DoesNotExist:
                pass

        return Response({
            "token": token.key,
            "user": {
                "codigo": u.codigo,
                "nombre": u.nombre,
                "apellido": u.apellido,
                "correo": u.correo,
                "sexo": u.sexo,
                "telefono": u.telefono,
                "estado": u.estado,
                "idrol": u.idrol_id,
                "rol": rol_obj,
            }
        }, status=status.HTTP_200_OK)


# CU02. Registrarse en el sistema
class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        data = request.data.copy()

        # Validaciones mínimas
        required = ["nombre", "apellido", "correo", "contrasena"]
        missing = [f for f in required if not data.get(f)]
        if missing:
            return Response(
                {"detail": f"Faltan campos requeridos: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ¿correo ya existe?
        if Usuario.objects.filter(correo=data["correo"]).exists():
            return Response(
                {"detail": "El correo ya está registrado."},
                status=status.HTTP_409_CONFLICT
            )

        # ---- Defaults solo si NO vienen desde el front ----
        # estado: si no viene o viene vacío => "pendiente"
        estado = (data.get("estado") or "").strip()
        if not estado:
            data["estado"] = "pendiente"

        # idrol: si no viene => 2
        if data.get("idrol") in [None, "", 0, "0"]:
            data["idrol"] = 2
        # (opcional) convertir a int si vino en string
        try:
            data["idrol"] = int(data["idrol"])
        except Exception:
            return Response({"detail": "idrol debe ser numérico."}, status=status.HTTP_400_BAD_REQUEST)

        # Crea Usuario con serializer general
        serializer = UsuarioSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        u = serializer.save()  # Usuario creado

        # Sincroniza auth.User (para emitir token y usar TokenAuth)
        dj_user, created = User.objects.get_or_create(
            username=u.correo,
            defaults={"email": u.correo, "first_name": u.nombre, "last_name": u.apellido},
        )
        dj_user.set_password(data["contrasena"])  # hash en Django
        dj_user.is_active = True
        dj_user.save()

        token, _ = Token.objects.get_or_create(user=dj_user)

        # Info de rol (si existe)
        rol_obj = None
        if u.idrol_id:
            try:
                r = Rol.objects.get(pk=u.idrol_id)
                rol_obj = {
                    "id": r.id,
                    "descripcion": r.descripcion,
                    "tipo": r.tipo,
                    "estado": r.estado,
                }
            except Rol.DoesNotExist:
                pass

        return Response({
            "token": token.key,
            "user": {
                "codigo": u.codigo,
                "nombre": u.nombre,
                "apellido": u.apellido,
                "correo": u.correo,
                "sexo": u.sexo,
                "telefono": u.telefono,
                "estado": u.estado,  # lo que mandó el front o "pendiente"
                "idrol": u.idrol_id,  # lo que mandó el front o 2
                "rol": rol_obj,
            }
        }, status=status.HTTP_201_CREATED)


# CU04. Cierre de sesion
class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        # Si el cliente envía Authorization: Token <...>, DRF pone el token en request.auth
        all_sessions = bool(request.data.get("all", False))

        if all_sessions:
            Token.objects.filter(user=request.user).delete()
            detail = "Sesiones cerradas en todos los dispositivos."
        else:
            # Borra SOLO el token de esta sesión
            try:
                if request.auth:
                    request.auth.delete()
            except Exception:
                # si ya estaba borrado/no válido, seguimos devolviendo 200 para idempotencia
                pass
            detail = "Sesión cerrada."

        return Response({"detail": detail}, status=status.HTTP_200_OK)


# Agregar estos ViewSets al final de api/views.py, después de LogoutView y antes de AIDetectionViewSet:

class ReconocimientoFacialViewSet(BaseModelViewSet):
    queryset = ReconocimientoFacial.objects.all().order_by('-fecha_deteccion')
    serializer_class = ReconocimientoFacialSerializer
    filterset_fields = ['codigo_usuario', 'es_residente', 'ubicacion_camara', 'estado', 'fecha_deteccion']
    search_fields = ['ubicacion_camara', 'estado']
    ordering_fields = ['id', 'fecha_deteccion', 'confianza']


class DeteccionPlacaViewSet(BaseModelViewSet):
    queryset = DeteccionPlaca.objects.all().order_by('-fecha_deteccion')
    serializer_class = DeteccionPlacaSerializer
    filterset_fields = ['placa_detectada', 'vehiculo', 'es_autorizado', 'ubicacion_camara', 'tipo_acceso',
                        'fecha_deteccion']
    search_fields = ['placa_detectada', 'ubicacion_camara', 'tipo_acceso']
    ordering_fields = ['id', 'fecha_deteccion', 'confianza']


class PerfilFacialViewSet(BaseModelViewSet):
    queryset = PerfilFacial.objects.all().order_by('-fecha_registro')
    serializer_class = PerfilFacialSerializer
    filterset_fields = ['codigo_usuario', 'activo', 'fecha_registro']
    search_fields = ['codigo_usuario__nombre', 'codigo_usuario__apellido', 'codigo_usuario__correo']
    ordering_fields = ['id', 'fecha_registro']


class ReporteSeguridadViewSet(BaseModelViewSet):
    queryset = ReporteSeguridad.objects.all().order_by('-fecha_evento')
    serializer_class = ReporteSeguridadSerializer
    filterset_fields = ['tipo_evento', 'nivel_alerta', 'revisado', 'revisor', 'fecha_evento']
    search_fields = ['descripcion', 'tipo_evento', 'nivel_alerta']
    ordering_fields = ['id', 'fecha_evento', 'nivel_alerta']


logger = logging.getLogger(__name__)

from .services.ai_detection import FacialRecognitionService, PlateDetectionService

try:
    from .services.ai_detection import FacialRecognitionService, PlateDetectionService
except ImportError:
    FacialRecognitionService = None
    PlateDetectionService = None

# Solo definimos la clase AIDetectionViewSet si los servicios se importaron correctamente
if FacialRecognitionService and PlateDetectionService:
    class AIDetectionViewSet(viewsets.ViewSet):
        permission_classes = [permissions.IsAuthenticated]
        parser_classes = [MultiPartParser, FormParser]

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.storage_service = SupabaseStorageService()
            self.facial_service = FacialRecognitionService()
            self.plate_service = PlateDetectionService()

        # ... (Aquí va TODO el código de la clase AIDetectionViewSet, sin cambios en su interior)
        # ... (Desde @action(detail=False, methods=['post']) def recognize_face...)
        # ... (Hasta el final de la clase con la función detection_stats)

        # ============= RECONOCIMIENTO FACIAL =============
        @action(detail=False, methods=['post'])
        def recognize_face(self, request):
            # 1. Obtener la imagen (esto no cambia)
            image_file = request.FILES.get('image')
            camera_location = request.data.get('camera_location', 'Principal')

            if not image_file:
                return Response(
                    {'error': 'La imagen es requerida como archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 2. Enviar la imagen al worker local
            try:
                if not AI_WORKER_URL:
                    # Fallback de seguridad si la URL no está configurada
                    logger.error("AI_WORKER_URL no está configurada en las variables de entorno.")
                    return Response({'error': 'El servicio de IA no está configurado'}, status=503)

                files = {'image': (image_file.name, image_file.read(), image_file.content_type)}

                # El endpoint en tu worker.py
                worker_endpoint = f"{AI_WORKER_URL}/recognize_face"

                # Hacemos la petición a tu PC
                response = requests.post(worker_endpoint, files=files, timeout=60)  # Timeout de 60 segundos
                response.raise_for_status()  # Lanza un error si la respuesta no es 2xx

                # El resultado que viene desde tu PC
                result = response.json()

            except requests.exceptions.RequestException as e:
                logger.error(f"Error contactando al worker de IA: {e}")
                return Response({'error': 'El servicio de IA no está disponible o tardó demasiado en responder'},
                                status=503)

            # 3. Usar el resultado para crear el Reporte de Seguridad (esta lógica se queda aquí)
            try:
                if not result.get('is_resident'):
                    ReporteSeguridad.objects.create(
                        tipo_evento='intruso_detectado',
                        reconocimiento_facial_id=result.get('id'),
                        descripcion=f"Persona no identificada detectada en {camera_location}",
                        nivel_alerta='alto'
                    )

                logger.info(f"Reconocimiento completado por el worker - residente: {result.get('is_resident')}")
                return Response(result, status=status.HTTP_200_OK)

            except Exception as e:
                # Este error ocurriría si hay un problema con la base de datos al guardar el reporte
                logger.error(f"Error guardando el resultado del reconocimiento: {e}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                return Response(
                    {'error': 'Error interno del servidor al procesar el resultado'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # ============= REGISTRO DE PERFILES FACIALES =============
        @action(detail=False, methods=['post'])
        def register_profile(self, request):
            """Registra un nuevo perfil facial para un usuario específico"""
            try:
                user_id = request.data.get('user_id')
                image_file = request.FILES.get('image')

                if not user_id or not image_file:
                    return Response({
                        'success': False,
                        'error': 'user_id e imagen son requeridos'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Verificar que el usuario existe
                try:
                    usuario = Usuario.objects.get(codigo=user_id)
                except Usuario.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Usuario no encontrado'
                    }, status=status.HTTP_404_NOT_FOUND)

                # USAR EL NUEVO MÉTODO QUE MANEJA ARCHIVOS DIRECTAMENTE
                success = self.facial_service.register_face_from_file(int(user_id), image_file)

                if success:
                    # Obtener el perfil creado
                    perfil = PerfilFacial.objects.get(codigo_usuario=usuario)

                    return Response({
                        'success': True,
                        'message': f'Perfil facial registrado exitosamente para {usuario.nombre} {usuario.apellido}',
                        'profile_id': perfil.id,
                        'user': {
                            'codigo': usuario.codigo,
                            'nombre': usuario.nombre,
                            'apellido': usuario.apellido,
                            'correo': usuario.correo
                        },
                        'image_url': perfil.imagen_url
                    }, status=status.HTTP_201_CREATED)
                else:
                    return Response({
                        'success': False,
                        'error': 'No se pudo registrar el perfil facial'
                    }, status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                logger.error(f"Error registrando perfil facial: {str(e)}")
                return Response({
                    'success': False,
                    'error': f'Error interno del servidor: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        @action(detail=False, methods=['post'])
        def register_current_user(self, request):
            """Registra perfil facial del usuario autenticado actual"""
            try:
                # Obtener imagen del FormData
                image_file = request.FILES.get('image')

                logger.info(f"Datos recibidos - image_file: {'Si' if image_file else 'No'}")

                if not image_file:
                    return Response({
                        'success': False,
                        'error': 'La imagen es requerida como archivo'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Obtener usuario autenticado desde token
                try:
                    usuario = Usuario.objects.get(correo=request.user.email)
                    logger.info(f"Usuario encontrado: {usuario.nombre} {usuario.apellido}")
                except Usuario.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Usuario no encontrado en el sistema'
                    }, status=status.HTTP_404_NOT_FOUND)

                # Verificar si ya tiene un perfil registrado
                existing_profile = PerfilFacial.objects.filter(codigo_usuario=usuario, activo=True).first()
                if existing_profile:
                    logger.info(f"Usuario ya tiene perfil facial: {existing_profile.id}")
                    return Response({
                        'success': False,
                        'error': 'Ya tienes un perfil facial registrado. Elimínalo primero si quieres crear uno nuevo.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # USAR EL NUEVO MÉTODO QUE MANEJA ARCHIVOS DIRECTAMENTE
                success = self.facial_service.register_face_from_file(usuario.codigo, image_file)

                if success:
                    # Obtener el perfil creado por el servicio
                    profile = PerfilFacial.objects.get(codigo_usuario=usuario, activo=True)

                    logger.info(f"Perfil facial creado exitosamente: {profile.id}")

                    return Response({
                        'success': True,
                        'message': f'Perfil facial registrado exitosamente para {usuario.nombre}',
                        'profile_id': profile.id,
                        'user': {
                            'codigo': usuario.codigo,
                            'nombre': usuario.nombre,
                            'apellido': usuario.apellido,
                            'correo': usuario.correo
                        },
                        'image_url': profile.imagen_url
                    }, status=status.HTTP_201_CREATED)
                else:
                    return Response({
                        'success': False,
                        'error': 'No se pudo registrar el perfil facial. Verifica que la imagen contenga una cara visible.'
                    }, status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                logger.error(f"Error registrando perfil del usuario actual: {str(e)}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                return Response({
                    'success': False,
                    'error': 'Error interno del servidor'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        @action(detail=False, methods=['get'])
        def list_profiles(self, request):
            """Lista todos los perfiles faciales registrados"""
            try:
                profiles = PerfilFacial.objects.filter(activo=True).select_related('codigo_usuario')

                profiles_data = []
                for profile in profiles:
                    profiles_data.append({
                        'id': profile.id,
                        'user_id': profile.codigo_usuario.codigo,
                        'user_name': f"{profile.codigo_usuario.nombre} {profile.codigo_usuario.apellido}",
                        'user_email': profile.codigo_usuario.correo,
                        'image_url': profile.imagen_url,
                        'fecha_registro': profile.fecha_registro.isoformat(),
                        'activo': profile.activo
                    })

                return Response({
                    'success': True,
                    'profiles': profiles_data,
                    'count': len(profiles_data)
                }, status=status.HTTP_200_OK)

            except Exception as e:
                logger.error(f"Error listando perfiles: {str(e)}")
                return Response({
                    'success': False,
                    'error': 'Error interno del servidor',
                    'profiles': [],
                    'count': 0
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Y la función delete_profile:

            # En api/views.py, dentro de la clase AIDetectionViewSet

            @action(detail=True, methods=['delete'])
            def delete_profile(self, request, pk=None):
                """Elimina un perfil facial"""
                try:
                    perfil = PerfilFacial.objects.get(id=pk)
                    user_name = f"{perfil.codigo_usuario.nombre} {perfil.codigo_usuario.apellido}"

                    # Eliminar imagen de Supabase Storage
                    if perfil.imagen_path:
                        delete_success = self.storage_service.delete_file(perfil.imagen_path)
                        if not delete_success:
                            logger.warning(f"No se pudo eliminar la imagen {perfil.imagen_path} de Supabase.")

                    # Eliminar registro de la base de datos
                    perfil.delete()

                    # --- LÍNEA CORREGIDA ---
                    # Ya no intentamos recargar los perfiles aquí.

                    return Response({
                        'success': True,
                        'message': f'Perfil facial de {user_name} eliminado exitosamente'
                    }, status=status.HTTP_200_OK)  # Cambiado a 200 OK para devolver un mensaje

                except PerfilFacial.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'Perfil no encontrado'
                    }, status=status.HTTP_404_NOT_FOUND)
                except Exception as e:
                    logger.error(f"Error eliminando perfil: {str(e)}")
                    return Response({
                        'success': False,
                        'error': 'Ocurrió un error interno al eliminar el perfil.'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        @action(detail=False, methods=['get'])
        def list_profiles(self, request):
            """Lista todos los perfiles faciales registrados"""
            try:
                perfiles = PerfilFacial.objects.select_related('codigo_usuario').filter(activo=True)

                profiles_data = []
                for perfil in perfiles:
                    profiles_data.append({
                        'id': perfil.id,
                        'user_id': perfil.codigo_usuario.codigo,
                        'user_name': f"{perfil.codigo_usuario.nombre} {perfil.codigo_usuario.apellido}",
                        'user_email': perfil.codigo_usuario.correo,
                        'image_url': perfil.imagen_url,
                        'fecha_registro': perfil.fecha_registro.isoformat(),
                        'activo': perfil.activo
                    })

                return Response({
                    'success': True,
                    'profiles': profiles_data,
                    'count': len(profiles_data)
                }, status=status.HTTP_200_OK)

            except Exception as e:
                logger.error(f"Error obteniendo perfiles: {str(e)}")
                return Response({
                    'success': False,
                    'error': 'Error obteniendo perfiles'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        @action(detail=True, methods=['delete'])
        def delete_profile(self, request, pk=None):
            """Elimina un perfil facial"""
            try:
                perfil = PerfilFacial.objects.get(id=pk)

                # Eliminar imagen de Supabase Storage
                if perfil.imagen_path:
                    self.storage_service.delete_file(perfil.imagen_path)

                # Eliminar registro
                user_name = f"{perfil.codigo_usuario.nombre} {perfil.codigo_usuario.apellido}"
                perfil.delete()

                # Recargar perfiles faciales en el servicio
                self.facial_service.load_known_faces()

                return Response({
                    'success': True,
                    'message': f'Perfil facial de {user_name} eliminado exitosamente'
                }, status=status.HTTP_200_OK)

            except PerfilFacial.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Perfil no encontrado'
                }, status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                logger.error(f"Error eliminando perfil: {str(e)}")
                return Response({
                    'success': False,
                    'error': 'Error eliminando perfil'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # ============= DETECCIÓN DE PLACAS =============
        @action(detail=False, methods=['post'])
        def detect_plate(self, request):
            try:
                # Obtener datos del FormData
                image_file = request.FILES.get('image')
                camera_location = request.data.get('camera_location', 'Estacionamiento')
                access_type = request.data.get('access_type', 'entrada')

                if not image_file:
                    return Response(
                        {'error': 'La imagen es requerida como archivo'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                logger.info(f"Procesando detección de placa - cámara: {camera_location}, tipo: {access_type}")

                # USAR EL NUEVO MÉTODO QUE MANEJA ARCHIVOS DIRECTAMENTE
                result = self.plate_service.detect_plate_from_file(
                    image_file, camera_location, access_type
                )

                # Crear reporte de seguridad si la placa no está autorizada
                if result['plate'] and not result['is_authorized']:
                    ReporteSeguridad.objects.create(
                        tipo_evento='placa_no_autorizada',
                        deteccion_placa_id=result['id'],
                        descripcion=f"Placa no autorizada detectada: {result['plate']} en {camera_location}",
                        nivel_alerta='medio'
                    )

                logger.info(f"Detección completada - placa: {result.get('plate', 'No detectada')}")
                return Response(result, status=status.HTTP_200_OK)

            except Exception as e:
                logger.error(f"Error en detección de placa: {e}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                return Response(
                    {'error': 'Error interno del servidor'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # ============= ESTADÍSTICAS =============
        @action(detail=False, methods=['get'])
        def detection_stats(self, request):
            try:
                now = timezone.now()
                last_24h = now - timedelta(hours=24)
                last_week = now - timedelta(days=7)

                stats = {
                    'facial_recognitions_today': ReconocimientoFacial.objects.filter(
                        fecha_deteccion__gte=last_24h
                    ).count(),
                    'plate_detections_today': DeteccionPlaca.objects.filter(
                        fecha_deteccion__gte=last_24h
                    ).count(),
                    'residents_detected_today': ReconocimientoFacial.objects.filter(
                        fecha_deteccion__gte=last_24h,
                        es_residente=True
                    ).count(),
                    'unauthorized_plates_today': DeteccionPlaca.objects.filter(
                        fecha_deteccion__gte=last_24h,
                        es_autorizado=False
                    ).count(),
                    'security_alerts_week': ReporteSeguridad.objects.filter(
                        fecha_evento__gte=last_week,
                        nivel_alerta__in=['alto', 'critico']
                    ).count(),
                    'registered_faces': PerfilFacial.objects.filter(activo=True).count()
                }

                return Response(stats, status=status.HTTP_200_OK)

            except Exception as e:
                logger.error(f"Error obteniendo estadísticas: {e}")
                return Response(
                    {'error': 'Error interno del servidor'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

else:
    # Si la importación falla, definimos AIDetectionViewSet como None
    AIDetectionViewSet = None


# -------- Helpers ----------
def _month_range(yyyy_mm: str):
    """Devuelve (primer_día, último_día) para un 'YYYY-MM'. Si es inválido, usa el mes actual."""
    try:
        y, m = map(int, yyyy_mm.split("-"))
        first = date(y, m, 1)
    except Exception:
        today = date.today()
        y, m = today.year, today.month
        first = date(y, m, 1)

    if m == 12:
        last = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(y, m + 1, 1) - timedelta(days=1)
    return first, last


def _bitacora(request, accion: str):
    try:
        u = Usuario.objects.get(correo=request.user.email)
        Bitacora.objects.create(
            codigo_usuario=u,
            accion=accion,
            fecha=timezone.now().date(),
            hora=timezone.now().time(),
            ip=request.META.get("HTTP_X_FORWARDED_FOR", "").split(",")[0] or request.META.get(
                "REMOTE_ADDR") or "0.0.0.0",
        )
    except Usuario.DoesNotExist:
        pass


# -------- Endpoint: Estado de Cuenta ----------
class EstadoCuentaView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = Usuario.objects.get(correo=request.user.email)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no encontrado."}, status=404)

        mes = request.query_params.get("mes") or date.today().strftime("%Y-%m")
        desde, hasta = _month_range(mes)

        # --- LÓGICA DE CARGOS (SIN CAMBIOS EN CÓMO SE OBTIENEN) ---
        pagos_realizados_en_mes = Factura.objects.filter(
            codigo_usuario=user,
            fecha__range=(desde, hasta),
            estado="pagado",
            id_pago__isnull=False
        ).values_list('id_pago_id', flat=True)

        cargos = []
        pagos_catalogo_qs = Pagos.objects.filter(
            tipo__in=['Mantenimiento', 'Extraordinaria']
        )

        for p in pagos_catalogo_qs:
            cargos.append({
                "id": p.id,
                "tipo": p.tipo,
                "descripcion": p.descripcion,
                "monto": p.monto,
                "origen": "pago",
                "fecha": None,
                "pagado": p.id in pagos_realizados_en_mes
            })

        # ... (La lógica de multas no cambia)

        # --- LÓGICA DE CÁLCULOS CORREGIDA ---

        # 1. Total Cargos del Mes: Suma de TODOS los cargos del catálogo (sin importar si están pagados o no)
        total_cargos_del_mes = sum(Decimal(c["monto"] or 0) for c in cargos if c['origen'] != 'multa')

        # 2. Total Pagos del Mes: (Esta parte ya estaba bien)
        pagos_realizados_ser = PagoRealizadoSerializer(
            Factura.objects.filter(codigo_usuario=user, fecha__range=(desde, hasta), estado="pagado"),
            many=True
        ).data
        total_pagos_del_mes = sum(Decimal(p['monto']) for p in pagos_realizados_ser)

        # 3. Saldo Pendiente: La diferencia entre el total de cargos y el total de pagos
        saldo_pendiente = total_cargos_del_mes - total_pagos_del_mes

        # --- FIN DE LA CORRECCIÓN ---

        payload = {
            "mes": mes,
            "cargos": cargos,
            "pagos": pagos_realizados_ser,
            # Enviamos los nuevos valores calculados
            "totales": {
                "cargos": f"{total_cargos_del_mes:.2f}",
                "pagos": f"{total_pagos_del_mes:.2f}",
                "saldo": f"{saldo_pendiente:.2f}",
            },
        }
        return Response(payload, status=200)


# -------- Endpoint: PDF Comprobante ----------
# En api/views.py

class ComprobantePDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        try:
            user = Usuario.objects.get(correo=request.user.email)
            factura = (
                Factura.objects
                .select_related("id_pago", "codigo_usuario")
                .get(id=pk, codigo_usuario=user)
            )
        except (Usuario.DoesNotExist, Factura.DoesNotExist):
            raise Http404()

        buff = BytesIO()
        c = canvas.Canvas(buff, pagesize=A4)
        w, h = A4

        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h - 60, "Smart Condominium - Comprobante de Pago")

        c.setFont("Helvetica", 11)
        y = h - 110

        # --- CÓDIGO MEJORADO Y MÁS SEGURO ---
        # Nos aseguramos de que todos los datos sean texto antes de dibujarlos
        rows = [
            ("N° Comprobante", str(factura.id)),
            ("Fecha", factura.fecha.strftime("%Y-%m-%d") if factura.fecha else "N/A"),
            ("Hora", factura.hora.strftime("%H:%M:%S") if factura.hora else "N/A"),
            ("Usuario", f"{factura.codigo_usuario.nombre} {factura.codigo_usuario.apellido}"),
            ("Correo", factura.codigo_usuario.correo),
            ("Concepto", factura.id_pago.descripcion if factura.id_pago else "Concepto no disponible"),
            ("Tipo de Pago", str(factura.tipo_pago)),
            ("Monto (Bs.)",
             f"{factura.id_pago.monto:.2f}" if factura.id_pago and factura.id_pago.monto is not None else "0.00"),
            ("Estado", str(factura.estado)),
        ]

        for label, value in rows:
            c.drawString(40, y, f"{label}: {value}")
            y -= 20

        c.line(40, y - 10, w - 40, y - 10)
        c.drawString(40, y - 30, "Gracias por su pago.")
        c.showPage()
        c.save()
        buff.seek(0)

        _bitacora(request, f"Descarga comprobante #{factura.id}")

        return FileResponse(buff, as_attachment=True, filename=f"comprobante_{factura.id}.pdf")


# En api/views.py

class SolicitudMantenimientoViewSet(BaseModelViewSet):
    serializer_class = SolicitudMantenimientoSerializer

    def get_queryset(self):
        """
        Admins ven todas las solicitudes.
        Residentes ven solo las suyas.
        """
        try:
            usuario = Usuario.objects.get(correo=self.request.user.email)
            if usuario.idrol and usuario.idrol.tipo == 'admin':
                return SolicitudMantenimiento.objects.all().select_related('codigo_usuario', 'codigo_propiedad', 'id_pago')
            return SolicitudMantenimiento.objects.filter(codigo_usuario=usuario).select_related('codigo_usuario', 'codigo_propiedad', 'id_pago')
        except Usuario.DoesNotExist:
            return SolicitudMantenimiento.objects.none()

    def perform_create(self, serializer):
        """
        Asigna automáticamente el usuario y su propiedad al crear la solicitud.
        """
        try:
            usuario = Usuario.objects.get(correo=self.request.user.email)
            pertenencia_activa = Pertenece.objects.filter(codigo_usuario=usuario, fecha_fin__isnull=True).first()
            if not pertenencia_activa:
                raise serializers.ValidationError("No tienes una propiedad activa asignada para crear una solicitud.")
            serializer.save(codigo_usuario=usuario, codigo_propiedad=pertenencia_activa.codigo_propiedad)
        except Usuario.DoesNotExist:
            raise serializers.ValidationError("Usuario no encontrado.")

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin])
    def update_status(self, request, pk=None):
        solicitud = self.get_object()
        nuevo_estado = request.data.get('estado')

        if nuevo_estado not in ['En Progreso', 'Completada', 'Cancelada']:
            return Response({'error': 'Estado no válido.'}, status=status.HTTP_400_BAD_REQUEST)

        solicitud.estado = nuevo_estado
        solicitud.save()

        if nuevo_estado == 'Completada' and solicitud.id_pago:
            Factura.objects.create(
                codigo_usuario=solicitud.codigo_usuario,
                id_pago=solicitud.id_pago,
                fecha=timezone.now().date(),
                hora=timezone.now().time(),
                tipo_pago='Servicio Solicitado',
                estado='pendiente'
            )
        return Response(self.get_serializer(solicitud).data)

    # --- ¡NUEVA FUNCIÓN AÑADIDA! ---
    @action(detail=False, methods=['get'], url_path='servicios-disponibles')
    def list_services(self, request):
        """
        Devuelve una lista de todos los items del catálogo de Pagos
        que están marcados como 'Servicio'.
        """
        servicios = Pagos.objects.filter(tipo='Servicio')
        serializer = PagoSerializer(servicios, many=True)
        return Response(serializer.data)


import traceback


# api/views.py
class ReporteUsoAreasComunesView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def _get_report_data(self, fecha_inicio_str, fecha_fin_str):
        reservas_query = Reserva.objects.filter(
            fecha__range=[fecha_inicio_str, fecha_fin_str],
            estado='Aprobada'
        ).select_related('id_area_c')

        reporte_data = list(reservas_query.values(
            'id_area_c__descripcion'
        ).annotate(
            cantidad_reservas=Count('id')
        ).order_by('-cantidad_reservas'))

        total_reservas = reservas_query.count()
        area_mas_usada = reporte_data[0] if reporte_data else None
        area_menos_usada = reporte_data[-1] if len(reporte_data) > 1 else None

        return {
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_fin_str,
            'total_reservas': total_reservas,
            'area_mas_usada': area_mas_usada,
            'area_menos_usada': area_menos_usada,
            'detalle_por_area': reporte_data
        }

    def _generate_pdf_response(self, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        # ... (El resto de la lógica para generar el PDF no necesita cambios, ya estaba bien)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Reporte de Uso de Áreas Comunes", styles['h1']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Periodo: {data['fecha_inicio']} al {data['fecha_fin']}", styles['h3']))
        elements.append(Spacer(1, 24))

        table_data = [["Área Común", "Cantidad de Reservas"]]
        if data['detalle_por_area']:
            for item in data['detalle_por_area']:
                table_data.append([str(item['id_area_c__descripcion']), str(item['cantidad_reservas'])])
        else:
            table_data.append(["No se encontraron reservas en este periodo.", ""])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_areas_comunes.pdf"'
        return response

    def _generate_excel_response(self, data):
        # ... (La lógica para generar Excel tampoco necesita cambios)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Uso de Áreas Comunes"
        ws['A1'] = "Reporte de Uso de Áreas Comunes"
        ws.merge_cells('A1:B1')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.append([])
        ws.append(["Periodo", f"{data['fecha_inicio']} a {data['fecha_fin']}"])
        ws.append([])
        headers = ["Área Común", "Cantidad de Reservas"]
        ws.append(headers)
        for cell in ws[5]:
            cell.font = Font(bold=True)
        if data['detalle_por_area']:
            for item in data['detalle_por_area']:
                ws.append([item['id_area_c__descripcion'], item['cantidad_reservas']])
        else:
            ws.append(["No se encontraron reservas.", ""])
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_areas_comunes.xlsx"'
        return response

    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        # --- CAMBIO IMPORTANTE: Usamos 'export' en lugar de 'format' ---
        export_param = request.query_params.get('export')

        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Las fechas de inicio y fin son obligatorias.'}, status=400)

        try:
            report_data = self._get_report_data(fecha_inicio, fecha_fin)

            if export_param == 'pdf':
                return self._generate_pdf_response(report_data)
            elif export_param == 'xlsx':
                return self._generate_excel_response(report_data)
            else:
                # Si no se pide exportar, se devuelven los datos para la tabla en pantalla
                return Response(report_data)

        except Exception as e:
            return Response({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)

# Pega esto al final de api/views.py

from django.http import JsonResponse


def test_view(request):
    """Una vista de prueba muy simple."""
    return JsonResponse({"mensaje": "¡La URL de prueba funciona!"})


class MantenimientoPreventivoViewSet(BaseModelViewSet):
    """
    Gestiona la programación de mantenimientos preventivos.
    Solo accesible para administradores.
    """
    queryset = MantenimientoPreventivo.objects.all()
    serializer_class = MantenimientoPreventivoSerializer
    permission_classes = [IsAdmin]  # Solo los admins pueden gestionar esto
    filterset_fields = ['estado', 'id_tarea', 'proxima_fecha']
    search_fields = ['id_tarea__descripcion', 'descripcion_adicional']
    ordering_fields = ['proxima_fecha', 'fecha_inicio', 'frecuencia_dias']


from django.db import transaction


# En api/views.py, reemplaza la clase PagarCuotaView

class PagarCuotaView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        id_pago = request.data.get('id_pago')
        mes = request.data.get('mes')

        if not id_pago or not mes:
            return Response({'error': 'Se requiere el id del pago y el mes.'}, status=400)

        try:
            with transaction.atomic():
                usuario = Usuario.objects.get(correo=request.user.email)
                pago_info = Pagos.objects.get(id=id_pago)

                # Validar si ya existe una factura pagada (la lógica de antes)
                fecha_inicio_mes, fecha_fin_mes = _month_range(mes)
                if Factura.objects.filter(
                        codigo_usuario=usuario,
                        id_pago=pago_info,
                        fecha__range=[fecha_inicio_mes, fecha_fin_mes],
                        estado='pagado'  # <-- Solo nos importa si ya está pagada
                ).exists():
                    return Response({'error': 'Esta cuota ya ha sido pagada para el mes seleccionado.'}, status=400)

                # 1. Creamos la factura pero como PENDIENTE
                factura_pendiente = Factura.objects.create(
                    codigo_usuario=usuario,
                    id_pago=pago_info,
                    fecha=timezone.now().date(),
                    hora=timezone.now().time(),
                    tipo_pago='Stripe (Pendiente)',  # Estado inicial
                    estado='pendiente'  # <-- Estado inicial
                )

                # 2. Creamos la sesión de Checkout de Stripe
                FRONTEND_URL = settings.FRONTEND_URL
                session = stripe.checkout.Session.create(
                    payment_method_types=['card'],
                    line_items=[{
                        'price_data': {
                            'currency': 'bob',
                            'product_data': {
                                'name': f"{pago_info.descripcion} (Mes: {mes})",
                            },
                            'unit_amount': int(pago_info.monto * 100),  # Stripe usa centavos
                        },
                        'quantity': 1,
                    }],
                    mode='payment',
                    # URLs a las que Stripe redirigirá al usuario
                    success_url=f'{FRONTEND_URL}/pago-exitoso?session_id={{CHECKOUT_SESSION_ID}}',
                    cancel_url=f'{FRONTEND_URL}/estado-cuenta',
                    # Guardamos el ID de nuestra factura para saber qué actualizar después
                    client_reference_id=factura_pendiente.id
                )

                # 3. Devolvemos el ID de la sesión de Stripe al frontend
                return Response({'sessionId': session.id})

        except Exception as e:
            return Response({'error': str(e)}, status=500)


from rest_framework.permissions import AllowAny  # <-- Importa AllowAny


class StripeWebhookView(APIView):
    """
    Escucha las notificaciones de Stripe para confirmar pagos.
    """
    permission_classes = [AllowAny]  # <-- Debe ser pública para que Stripe pueda llamarla

    def post(self, request):
        payload = request.body
        sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
        webhook_secret = os.getenv('STRIPE_WEBHOOK_SECRET')

        if not webhook_secret:
            return Response({'error': 'Webhook secret no configurado.'}, status=500)

        try:
            event = stripe.Webhook.construct_event(
                payload, sig_header, webhook_secret
            )
        except ValueError as e:
            return Response({'error': 'Payload inválido'}, status=400)
        except stripe.error.SignatureVerificationError as e:
            return Response({'error': 'Firma inválida'}, status=400)

        # Escuchamos el evento 'checkout.session.completed'
        if event['type'] == 'checkout.session.completed':
            session = event['data']['object']
            factura_id = session.get('client_reference_id')

            try:
                # ¡Pago exitoso! Actualizamos nuestra base de datos
                with transaction.atomic():
                    factura = Factura.objects.get(id=factura_id, estado='pendiente')
                    factura.estado = 'pagado'
                    factura.tipo_pago = 'Stripe (Real)'
                    factura.save()

                    # Y creamos los registros de finanzas y bitácora
                    Finanzas.objects.create(
                        tipo='Ingreso',
                        descripcion=f"Pago (Stripe) de '{factura.id_pago.descripcion}' por {factura.codigo_usuario.nombre}",
                        monto=factura.id_pago.monto,
                        fecha=timezone.now().date(),
                        origen='Pago Residente',
                        id_factura=factura
                    )

                    # No podemos llamar a _bitacora aquí porque no tenemos el 'request' del usuario
                    # Pero el pago ya está registrado.

            except Factura.DoesNotExist:
                return Response({'error': 'Factura no encontrada'}, status=404)
            except Exception as e:
                return Response({'error': str(e)}, status=500)

        return Response(status=200)

# smartcondominium-back/api/views.py

# smartcondominium-back/api/views.py

# ... (tus otras importaciones y clases de vistas)

class ReporteBitacoraView(APIView):
    """
    Genera un reporte de la bitácora del sistema entre un rango de fechas.
    Soporta formatos JSON, PDF y Excel.
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def _get_report_data(self, fecha_inicio, fecha_fin):
        """Obtiene y formatea los datos de la bitácora desde la base de datos."""
        query = Bitacora.objects.filter(
            fecha__range=[fecha_inicio, fecha_fin]
        ).select_related('codigo_usuario').order_by('-fecha', '-hora')

        detalle = []
        for entrada in query:
            detalle.append({
                "fecha": entrada.fecha.strftime('%Y-%m-%d'),
                "hora": entrada.hora.strftime('%H:%M:%S'),
                "usuario": f"{entrada.codigo_usuario.nombre} {entrada.codigo_usuario.apellido}" if entrada.codigo_usuario else "Sistema",
                "accion": entrada.accion,
                "ip": entrada.ip,
            })

        total_entradas = query.count()
        usuarios_activos = query.values('codigo_usuario').distinct().count()

        return {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_entradas": total_entradas,
            "total_usuarios_activos": usuarios_activos,
            "detalle": detalle
        }

    def _generate_pdf_response(self, data):
        """Genera la respuesta en formato PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Reporte de Bitácora del Sistema", styles['h1']))
        elements.append(Paragraph(f"Periodo: {data['fecha_inicio']} a {data['fecha_fin']}", styles['h3']))
        elements.append(Spacer(1, 24))

        table_data = [["Fecha", "Hora", "Usuario", "Acción", "Dirección IP"]]
        for item in data['detalle']:
            table_data.append([item['fecha'], item['hora'], item['usuario'], str(item['accion']), item['ip']])

        if not data['detalle']:
            table_data.append(["No hay entradas en el periodo seleccionado.", "", "", "", ""])

        table = Table(table_data, colWidths=[70, 60, 120, 350, 100])
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        if not data['detalle']:
            style.add('SPAN', (0, 1), (-1, 1))
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_bitacora.pdf"'
        return response

    def _generate_excel_response(self, data):
        """Genera la respuesta en formato Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Bitácora"
        ws['A1'] = "Reporte de Bitácora del Sistema"
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'] = f"Periodo: {data['fecha_inicio']} a {data['fecha_fin']}"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ["Fecha", "Hora", "Usuario", "Acción", "Dirección IP"]
        ws.append(headers)
        for cell in ws[4]:
            cell.font = Font(bold=True)

        for item in data['detalle']:
            ws.append([item['fecha'], item['hora'], item['usuario'], item['accion'], item['ip']])

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_bitacora.xlsx"'
        return response

    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        export_param = request.query_params.get('export')

        if not fecha_inicio or not fecha_fin:
            return Response({'error': 'Los parámetros `fecha_inicio` y `fecha_fin` son obligatorios.'}, status=400)

        try:
            report_data = self._get_report_data(fecha_inicio, fecha_fin)

            if export_param == 'pdf':
                return self._generate_pdf_response(report_data)
            elif export_param == 'xlsx':
                return self._generate_excel_response(report_data)
            else:
                return Response(report_data)

        except Exception as e:
            print(traceback.format_exc())
            return Response({'error': f'Ocurrió un error al generar el reporte: {str(e)}'}, status=500)

class HistorialPagosView(APIView):
    """
    Devuelve el historial completo de pagos para el usuario autenticado.
    Soporta exportación a PDF.
    """
    permission_classes = [IsAuthenticated]

    def _get_history_data(self, user):
        """Obtiene y serializa el historial de facturas pagadas."""
        historial_qs = Factura.objects.filter(
            codigo_usuario=user,
            estado='pagado'
        ).select_related('id_pago').order_by('-fecha', '-hora')

        serializer = PagoRealizadoSerializer(historial_qs, many=True)
        return serializer.data

    def _generate_pdf_history(self, history_data, user):
        """Genera un archivo PDF con el historial de pagos."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Historial de Pagos", styles['h1']))
        elements.append(Paragraph(f"Residente: {user.nombre} {user.apellido}", styles['h3']))
        elements.append(Spacer(1, 24))

        # --- CÓDIGO MEJORADO Y MÁS SEGURO ---
        # 1. Encabezados de la tabla
        table_data = [["Fecha", "Concepto", "Monto (Bs.)", "Tipo de Pago"]]

        # 2. Convertimos CADA dato a string antes de añadirlo a la tabla
        for item in history_data:
            table_data.append([
                str(item['fecha']),
                str(item['concepto']),
                str(item['monto']),
                str(item['tipo_pago'])
            ])

        if not history_data:
            table_data.append(["No se encontraron pagos registrados.", "", "", ""])

        table = Table(table_data, colWidths=[80, 250, 80, 100])
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ])
        if not history_data:
            style.add('SPAN', (0, 1), (-1, 1))
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historial_de_pagos.pdf"'
        return response

    def get(self, request):
        export_param = request.query_params.get('export')

        try:
            usuario = Usuario.objects.get(correo=request.user.email)
            history_data = self._get_history_data(usuario)

            if export_param == 'pdf':
                return self._generate_pdf_history(history_data, usuario)
            else:
                return Response(history_data)

        except Usuario.DoesNotExist:
            return Response({"error": "Usuario no encontrado."}, status=404)
        except Exception as e:
            print(traceback.format_exc())
            return Response({'error': f'Ocurrió un error: {str(e)}'}, status=500)


class MisNotificacionesView(APIView):
    """
    Devuelve las notificaciones para el usuario autenticado.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            usuario = Usuario.objects.get(correo=request.user.email)
            # Buscamos en la tabla Envio las notificaciones para este usuario
            envios = Envio.objects.filter(
                codigo_usuario=usuario
            ).select_related('id_notific').order_by('-fecha', '-hora')

            # Formateamos la respuesta
            notificaciones = []
            for envio in envios:
                notif = envio.id_notific
                notificaciones.append({
                    "id_envio": envio.id,
                    "tipo": notif.tipo,
                    "descripcion": notif.descripcion,
                    "fecha": f"{envio.fecha} {envio.hora}",
                    "estado": envio.estado
                })

            return Response(notificaciones)
        except Usuario.DoesNotExist:
            return Response({"error": "Usuario no encontrado."}, status=404)
